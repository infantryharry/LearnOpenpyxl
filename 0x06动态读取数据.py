#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   0x06动态读取数据.py
@Time    :   2021/12/11 22:22:09
@Desc    :   None
'''

# here put the import lib
import openpyxl

wb = openpyxl.load_workbook('./课件/测试2.xlsx')
ws = wb.worksheets[0]
wsMaxRow = ws.max_row
print(wsMaxRow)
wsMaxCol = ws.max_column
print(wsMaxCol)
