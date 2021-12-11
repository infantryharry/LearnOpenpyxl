#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   0x10按行列聚合.py
@Time    :   2021/12/11 22:59:52
@Desc    :   None
'''

# here put the import lib
import openpyxl

wb = openpyxl.load_workbook('./课件/练习2.xlsx')
ws = wb.worksheets[0]
# maxRow=ws.max_row
# maxCol=ws.max_column

# 按行聚合
print('按行聚合')
for row in list(ws.rows)[1:]:
    data = [i.value for i in row]
    print(data[0], data[1:])
print('\n')
# 按列聚合
print('按列聚合')
for col in list(ws.columns)[1:]:
    data = [i.value for i in col]
    print(data[0], data[1:])


wb.close
