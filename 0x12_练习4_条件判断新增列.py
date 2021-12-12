#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   0x12_练习4_条件判断新增列.py
@Time    :   2021/12/12 00:00:48
@Desc    :   None
'''

# here put the import lib
import openpyxl


wb = openpyxl.load_workbook(r'./课件/练习2.xlsx')
ws = wb['Sheet1']

ws['E1'] = '评价'
workRange = ws.iter_rows(min_row=2, min_col=2)
for rows in workRange:
    aaa = [i.value for i in rows][:-1]
    if sum(aaa) >= 270:
        rows[-1].value = '优秀'

wb.save(r'./课件/练习2_my.xlsx')
wb.close
