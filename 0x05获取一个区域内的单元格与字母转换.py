#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   0x005获取一个区域内的单元格与字母转换.py
@Time    :   2021/12/11 21:30:25
@Desc    :   None
'''

# here put the import lib
import openpyxl

wb = openpyxl.load_workbook('./课件/测试.xlsx')
ws = wb.worksheets[0]

# 打印表格里所有的值:方法一:sheet.vlues
sheetAllValues = list(ws.values)
# print(sheetAllValues)

# 方法二:ws.iter_rows(min_row=1, max_col=1,max_row=5, max_col=5 ):
# for row in ws.iter_rows(min_row=1, min_col=1, max_row=5, max_col=5):
#     for cell in row:
#         print(cell)

# 数字转字母 openpyxl.utils.get_column_letter(2)    # B
someLetter = openpyxl.utils.get_column_letter(2)
print(someLetter)

# 字母转数字openpyxl.utils.column_index_from_string('D')    # 4
someNum = openpyxl.utils.column_index_from_string('D')
print(someNum)

wb.close
