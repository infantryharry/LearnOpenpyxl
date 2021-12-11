#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   0x11_练习3_条件判断进行替换.py
@Time    :   2021/12/11 23:07:43
@Desc    :   None
'''

# here put the import lib
import openpyxl

wb = openpyxl.load_workbook('./课件/练习3.xlsx')
ws = wb.worksheets[0]
# maxRow=ws.max_row
# maxCol=ws.max_column
for row in ws.iter_rows(min_row=2, min_col=2):
    for i in row:
        if i.value is None:
            i.value = '缺考'
        elif i.value < 60:
            i.value = i.value + '(不及格)'

wb.save('./课件/练习3_my.xlsx')
wb.close
